import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from pathlib import Path

DATA_FILE = Path("data/sales_data.csv")
REPORT_File = Path("reports/sales_reports.xlsx")

#this reads from the CSV file
try:
    data = pd.read_csv(DATA_FILE)
except FileNotFoundError:
    print(f"Error: File not found at {DATA_FILE}")


#this will create the Workbook and the worksheet
wb = Workbook()
ws = wb.active
ws.title = "Sales Report"

#append Headers
ws.append(list(data.columns))

#writes the data into the rows
for row in data.itertuples(index=False):
    ws.append(row)

#styling the headers
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
  top=Side(style='thin'), bottom=Side(style='thin')
  )

for cell in ws[1]:
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')
    cell.fill = header_fill
    cell.border = thin_border

#adjusting the column widths 
for column_cells in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

#adding a summary row
totals_row_index = ws.max_row + 1
ws.cell(row=totals_row_index, column=1, value="Total").font = Font(bold=True)

# Sum numeric columns and write totals
for col_idx, col_name in enumerate(data.columns, start=1):
    if pd.api.types.is_numeric_dtype(data[col_name]):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Excel SUM formula for the column data (excluding header)
        sum_formula = f"=SUM({col_letter}2:{col_letter}{ws.max_row})"
        ws.cell(row=totals_row_index, column=col_idx, value=sum_formula).font = Font(bold=True)
        ws.cell(row=totals_row_index, column=col_idx).border = thin_border
    else:
        # For non-numeric columns, just leave the cell blank or could merge or style differently
        ws.cell(row=totals_row_index, column=col_idx).border = thin_border

# Style "Total" label cell with borders and alignment
total_label_cell = ws.cell(row=totals_row_index, column=1)
total_label_cell.alignment = Alignment(horizontal='center')
total_label_cell.border = thin_border
total_label_cell.fill = header_fill

wb.save(REPORT_File)
print(f"Report generated successfully: {REPORT_File}")